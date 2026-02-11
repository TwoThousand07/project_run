# from django.core.validators import URLValidator

from openpyxl import load_workbook, workbook


wb = load_workbook("upload_example.xlsx")

ws = wb.active

# for row in ws.iter_rows(min_row=1, max_col=5, values_only=True):
#     print(row)

# result_dict = {}

# for column in ws.iter_cols(values_only=True):
#     header = column[0]
#     values = [val for val in column[1:] if val is not None]
        
#     result_dict[header] = values

# for val in result_dict.values():
#     for i in range(len(val)):
#         print((val[i]))
#     print("\n")


result_list = []
rows = list(ws.iter_rows(values_only=True))

headers = rows[0]
data_rows = rows[1:]

for row in data_rows:
    data = dict(zip(headers, row))
    
    print(data, sep="\n")
    
# for row in result_list:
#     print(row, sep="\n")