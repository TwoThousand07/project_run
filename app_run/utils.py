def calculate_distance_between_two_positions(start: tuple, end: tuple) -> float:
    from geopy.distance import geodesic
    '''
    Docstring для calculate_distance_between_two_positions

    :param start_pos: Начальная точка
    :type start_pos: Position
    :param end_pos: Конец точки
    :type end_pos: Position
    :return: Возвращаем дистанцию между двумя точками
    :rtype: float
    '''

    return geodesic(start, end).km


def import_xlsx_from_file(xlsx_name) -> list[list]:
    from openpyxl import load_workbook
    from .serializers import CollectibleItemSerializer
    '''
        Получение xlsx файла и ее валидация, возвращает список непрошедших валидацию строк
    '''

    wb = load_workbook(xlsx_name)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    invalid_rows = []

    header = [str(header).strip().lower() for header in rows[0]]
    data_rows = rows[1:]


    # берем каждую строку отдельно, и валидируем
    # строки которые прошли валидацию сохраняются в БД
    # строки которые не прошли валидацию сохраняются в invalid_rows, и дальше обрабатываются в представлении
    for row in data_rows:
        if not any(row): continue
        
        raw_data = dict(zip(header, row))

        clean_data = {
            "name": raw_data.get("name"),
            "uid": raw_data.get("uid"),
            "value": raw_data.get("value"),
            "latitude": raw_data.get("latitude"),
            "longitude": raw_data.get("longitude"),
            "url": raw_data.get("url")
        }

        serializer = CollectibleItemSerializer(data=clean_data)

        if serializer.is_valid():
            serializer.save()
        else:
            invalid_rows.append(list(row))
    
    print(header)
    
    return invalid_rows
